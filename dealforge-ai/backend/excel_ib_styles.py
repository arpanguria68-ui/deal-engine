"""
Investment Banking Excel Style System
Extracted from WSO/CFI/Foresight professional templates.
Provides institutional-grade formatting for openpyxl workbooks.
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

# ═══════════════════════════════════════════
# IB COLOR PALETTE (from WSO 14-tab model)
# ═══════════════════════════════════════════
# Section header backgrounds
DARK_NAVY = "002060"
DARK_RED = "C00000"
MED_BLUE = "003366"
# Row fills
LIGHT_BLUE = "D9E2F3"
LIGHT_GREY = "F2F2F2"
WHITE = "FFFFFF"
LIGHT_GREEN = "E2EFDA"
LIGHT_YELLOW = "FFF2CC"
# Font colors
BLUE_INPUT = "0000FF"  # Hardcoded inputs  (editable)
BLACK_CALC = "000000"  # Calculated fields  (formulas)
RED_CHECK = "FF0000"  # Check / alert cells
GREEN_XREF = "006100"  # Cross-sheet references

# ═══════════════════════════════════════════
# FONT PRESETS
# ═══════════════════════════════════════════
TITLE_FONT = Font(name="Calibri", bold=True, size=16, color=WHITE)
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color=DARK_NAVY)
SECTION_FONT = Font(name="Calibri", bold=True, size=11, color=DARK_NAVY)
HDR_FONT = Font(name="Calibri", bold=True, size=10, color=WHITE)
LABEL_FONT = Font(name="Calibri", size=10, color=BLACK_CALC)
INPUT_FONT = Font(name="Calibri", size=10, color=BLUE_INPUT)  # Blue = editable
CALC_FONT = Font(name="Calibri", size=10, color=BLACK_CALC)  # Black = formula
CHECK_FONT = Font(name="Calibri", bold=True, size=10, color=RED_CHECK)
XREF_FONT = Font(name="Calibri", size=10, color=GREEN_XREF)
COVER_TITLE = Font(name="Calibri", bold=True, size=24, color=WHITE)
COVER_SUB = Font(name="Calibri", size=14, color=WHITE)
TOC_LINK = Font(name="Calibri", size=11, color=BLUE_INPUT, underline="single")
TINY_FONT = Font(name="Calibri", size=8, color="808080")

# ═══════════════════════════════════════════
# FILL PRESETS
# ═══════════════════════════════════════════
NAVY_FILL = PatternFill(start_color=DARK_NAVY, end_color=DARK_NAVY, fill_type="solid")
RED_FILL = PatternFill(start_color=DARK_RED, end_color=DARK_RED, fill_type="solid")
BLUE_FILL = PatternFill(start_color=MED_BLUE, end_color=MED_BLUE, fill_type="solid")
LBLUE_FILL = PatternFill(
    start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid"
)
LGREY_FILL = PatternFill(
    start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type="solid"
)
WHITE_FILL = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
GREEN_FILL = PatternFill(
    start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid"
)
YELLOW_FILL = PatternFill(
    start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type="solid"
)

# ═══════════════════════════════════════════
# BORDER PRESETS
# ═══════════════════════════════════════════
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
BOTTOM_BORDER = Border(bottom=Side(style="medium"))
TOP_BOTTOM_BORDER = Border(top=Side(style="medium"), bottom=Side(style="double"))

# ═══════════════════════════════════════════
# NUMBER FORMATS (from WSO templates)
# ═══════════════════════════════════════════
FMT_ACCOUNTING = '#,##0;(#,##0);"-"'  # Standard IB: neg in parens
FMT_ACCOUNTING1 = '#,##0.0;(#,##0.0);"-"'
FMT_PERCENT = "0.0%"
FMT_PERCENT2 = "0.00%"
FMT_DOLLAR = "$#,##0.00"
FMT_MULTIPLE = "0.0x"
FMT_YEAR = "0"
FMT_BILLIONS = '#,##0.0,,"B"'

# ═══════════════════════════════════════════
# HELPER FUNCTIONS
# ═══════════════════════════════════════════
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def style_header_row(ws, row, num_cols, font=HDR_FONT, fill=NAVY_FILL):
    """Apply IB header styling to a row."""
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_section_label(ws, row, col, text):
    """Write a bold navy section label."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = SECTION_FONT
    cell.border = BOTTOM_BORDER
    return cell


def style_input_cell(ws, row, col, value, fmt=FMT_ACCOUNTING):
    """Write a BLUE (editable input) cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = INPUT_FONT
    cell.number_format = fmt
    cell.border = THIN_BORDER
    return cell


def style_calc_cell(ws, row, col, value, fmt=FMT_ACCOUNTING):
    """Write a BLACK (calculated) cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = CALC_FONT
    cell.number_format = fmt
    cell.border = THIN_BORDER
    return cell


def style_total_row(ws, row, col, value, fmt=FMT_ACCOUNTING):
    """Write a bold total row with double-bottom border."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Calibri", bold=True, size=10, color=BLACK_CALC)
    cell.number_format = fmt
    cell.border = TOP_BOTTOM_BORDER
    return cell


def set_col_widths(ws, widths):
    """Set column widths from a dict of {letter: width}."""
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w


def alternate_row_shading(ws, start_row, end_row, num_cols):
    """Apply alternating light grey/white row shading."""
    for r in range(start_row, end_row + 1):
        fill = LGREY_FILL if (r - start_row) % 2 == 0 else WHITE_FILL
        for c in range(1, num_cols + 1):
            ws.cell(row=r, column=c).fill = fill
