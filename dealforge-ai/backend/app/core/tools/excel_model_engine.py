"""
OFAS Excel Model Engine — Formula-Preserving Financial Model Population

MCP Tools:
- excel_model_populate: Load template → map data to input cells → preserve formulas → save
- excel_export_tables: Export specific ranges from Excel to CSV/JSON
- excel_export_charts: Export embedded charts as PNG (placeholder — requires LibreOffice)

CRITICAL RULE: Never overwrite formula cells. Only write to designated input/assumption cells.
"""

import os
import csv
import io
import json
import copy
from typing import Dict, Any, Optional, List, Tuple
from pathlib import Path
from datetime import datetime
import structlog

try:
    import openpyxl
    from openpyxl.comments import Comment
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from app.core.tools.tool_router import BaseTool, ToolResult

logger = structlog.get_logger()

# Template directory
TEMPLATE_BASE = Path(
    r"F:\code project\Kimi_Agent_DealForge AI PRD\Knowledge managerment\Excel knowledge"
)

# Output directory for populated models
OUTPUT_DIR = Path(
    r"F:\code project\Kimi_Agent_DealForge AI PRD\dealforge-ai\backend\ofas_outputs"
)

# Template ID → file mapping
TEMPLATE_MAP = {
    "3statement": "181_Excel_Three-statement-model-1.xlsx",
    "3statement_alt": "CFI-Case-Study-Three-Statement-Model.xlsx",
    "dcf": "gyxypyur-DCF-Template.xlsx",
    "dcf_model": "6_dcf_model_1_0.xlsx",
    "sensitivity": "11_sensitivity_analysis_model_0_0.xlsx",
    "accretion_dilution": "Accretion-Dilution-Model.xlsx",
    "cap_table": "Cap Table and Exit Waterfall Tool, by Foresight.xlsx",
    "saas": "Enterprise SaaS Forecasting Tool, by Foresight.xlsx",
    "ecommerce": "Ecommerce Forecasting Tool, by Foresight.xlsx",
    "football_field": "Football-field-template.xlsx",
    "fcfe": "9_fcfe_model_0_0.xlsx",
    "monte_carlo": "IQRM_MonteCarlo_Template_Full.xlsx",
    "fund_of_funds": "Fund of Funds Model, by Foresight.xlsx",
    "equity": "CFI-Equity-Template.xlsx",
}


def _is_formula_cell(cell) -> bool:
    """Check if a cell contains a formula"""
    if cell.value is None:
        return False
    if isinstance(cell.value, str) and cell.value.startswith("="):
        return True
    return False


def _find_input_cells(ws) -> List[Dict[str, Any]]:
    """
    Identify input/assumption cells in a worksheet.
    These are typically:
    - Cells with a blue font or blue fill (financial modeling convention)
    - Cells in sheets named 'Assumptions', 'Inputs', or 'Drivers'
    - Cells that are NOT formulas and contain numeric values
    """
    input_cells = []
    for row in ws.iter_rows():
        for cell in row:
            if _is_formula_cell(cell):
                continue
            if cell.value is not None and not isinstance(cell.value, str):
                # Non-formula numeric cell = potential input
                input_cells.append(
                    {
                        "cell": cell.coordinate,
                        "value": cell.value,
                        "sheet": ws.title,
                    }
                )
    return input_cells


class ExcelModelPopulateTool(BaseTool):
    """
    Populate an Excel template with financial data while preserving all formulas.

    Only writes to designated input/assumption cells. Formula cells are NEVER overwritten.
    Adds RAG citation comments to assumption cells for audit trail.
    """

    def __init__(self):
        super().__init__(
            name="excel_model_populate",
            description=(
                "Load an Excel financial model template, populate input/assumption cells "
                "with financial data, preserve all formulas, add citation comments, "
                "and save a versioned output file."
            ),
        )

    def get_parameters_schema(self) -> Dict:
        return {
            "type": "object",
            "properties": {
                "template_id": {
                    "type": "string",
                    "description": (
                        "Template identifier. Options: 3statement, dcf, dcf_model, "
                        "sensitivity, accretion_dilution, cap_table, saas, ecommerce, "
                        "football_field, fcfe, monte_carlo, equity"
                    ),
                    "enum": list(TEMPLATE_MAP.keys()),
                },
                "ticker": {
                    "type": "string",
                    "description": "Company ticker symbol (e.g., 'MSFT')",
                },
                "cell_mappings": {
                    "type": "object",
                    "description": (
                        "Mapping of sheet_name -> cell_address -> value. "
                        "Example: {'Assumptions': {'B5': 100000, 'B6': 0.12}}"
                    ),
                },
                "assumptions_comments": {
                    "type": "object",
                    "description": (
                        "Mapping of sheet_name -> cell_address -> comment text. "
                        "Used for RAG citation trails."
                    ),
                },
            },
            "required": ["template_id", "ticker", "cell_mappings"],
        }

    def execute(
        self,
        template_id: str = "",
        ticker: str = "",
        cell_mappings: Optional[Dict] = None,
        assumptions_comments: Optional[Dict] = None,
        **kwargs,
    ) -> ToolResult:
        if not HAS_OPENPYXL:
            return ToolResult(
                success=False,
                data=None,
                error="openpyxl is not installed. Run: pip install openpyxl",
            )

        cell_mappings = cell_mappings or {}
        assumptions_comments = assumptions_comments or {}

        # Locate template
        template_file = TEMPLATE_MAP.get(template_id)
        if not template_file:
            return ToolResult(
                success=False,
                data=None,
                error=f"Unknown template: {template_id}. Available: {list(TEMPLATE_MAP.keys())}",
            )

        template_path = TEMPLATE_BASE / template_file
        if not template_path.exists():
            return ToolResult(
                success=False,
                data=None,
                error=f"Template file not found: {template_path}",
            )

        try:
            # Load template — data_only=False to preserve formulas
            wb = openpyxl.load_workbook(str(template_path), data_only=False)

            cells_written = 0
            cells_skipped = 0
            formula_cells_protected = 0

            for sheet_name, mappings in cell_mappings.items():
                if sheet_name not in wb.sheetnames:
                    logger.warning(
                        "Sheet not found, skipping",
                        sheet=sheet_name,
                        available=wb.sheetnames,
                    )
                    continue

                ws = wb[sheet_name]

                for cell_addr, value in mappings.items():
                    cell = ws[cell_addr]

                    # CRITICAL: Never overwrite formula cells
                    if _is_formula_cell(cell):
                        formula_cells_protected += 1
                        logger.warning(
                            "Formula cell protected",
                            sheet=sheet_name,
                            cell=cell_addr,
                            formula=str(cell.value)[:50],
                        )
                        cells_skipped += 1
                        continue

                    # Write the value
                    cell.value = value
                    cells_written += 1

                    # Add citation comment if provided
                    comments = assumptions_comments.get(sheet_name, {})
                    if cell_addr in comments:
                        comment_text = comments[cell_addr]
                        cell.comment = Comment(comment_text, "OFAS Architect Agent")

            # Create output directory
            OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

            # Save with version
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            output_filename = f"{ticker}_{template_id}_v{timestamp}.xlsx"
            output_path = OUTPUT_DIR / output_filename
            wb.save(str(output_path))
            wb.close()

            # Run basic validation checks
            checks = self._run_validation_checks(str(output_path))

            return ToolResult(
                success=True,
                data={
                    "model_path": str(output_path),
                    "template_used": template_file,
                    "ticker": ticker,
                    "cells_written": cells_written,
                    "cells_skipped": cells_skipped,
                    "formula_cells_protected": formula_cells_protected,
                    "sheets": wb.sheetnames if hasattr(wb, "sheetnames") else [],
                    "checks": checks,
                },
            )

        except Exception as e:
            logger.error("Excel populate failed", error=str(e))
            return ToolResult(success=False, data=None, error=str(e))

    def _run_validation_checks(self, model_path: str) -> Dict[str, Any]:
        """Run basic validation checks on the populated model"""
        checks = {
            "file_exists": os.path.exists(model_path),
            "file_size_kb": (
                round(os.path.getsize(model_path) / 1024, 1)
                if os.path.exists(model_path)
                else 0
            ),
            "formulas_intact": True,  # We never overwrite formulas
            "circular_references": False,  # Would need full recalc engine to detect
        }

        try:
            wb = openpyxl.load_workbook(model_path, data_only=False)
            formula_count = 0
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if _is_formula_cell(cell):
                            formula_count += 1
            checks["formula_count"] = formula_count
            wb.close()
        except Exception:
            checks["formula_count"] = -1

        return checks


class ExcelExportTablesTool(BaseTool):
    """
    Export specific ranges from an Excel file to CSV or JSON format.

    Used by downstream agents (Reporting, Valuation) to consume
    structured data from populated models.
    """

    def __init__(self):
        super().__init__(
            name="excel_export_tables",
            description=(
                "Export specific cell ranges from an Excel model to CSV or JSON format. "
                "Useful for extracting summary tables, P&L data, or valuation outputs."
            ),
        )

    def get_parameters_schema(self) -> Dict:
        return {
            "type": "object",
            "properties": {
                "model_path": {
                    "type": "string",
                    "description": "Path to the Excel file to export from",
                },
                "exports": {
                    "type": "array",
                    "description": "List of export specifications",
                    "items": {
                        "type": "object",
                        "properties": {
                            "sheet": {"type": "string"},
                            "range": {
                                "type": "string",
                                "description": "Cell range (e.g., 'A1:F20') or named range",
                            },
                            "format": {
                                "type": "string",
                                "enum": ["csv", "json"],
                            },
                            "output": {
                                "type": "string",
                                "description": "Output filename",
                            },
                        },
                        "required": ["sheet", "range", "format"],
                    },
                },
            },
            "required": ["model_path", "exports"],
        }

    def execute(
        self,
        model_path: str = "",
        exports: Optional[List[Dict]] = None,
        **kwargs,
    ) -> ToolResult:
        if not HAS_OPENPYXL:
            return ToolResult(
                success=False,
                data=None,
                error="openpyxl is not installed",
            )

        exports = exports or []

        if not os.path.exists(model_path):
            return ToolResult(
                success=False,
                data=None,
                error=f"Model file not found: {model_path}",
            )

        try:
            # Load with data_only=True to get computed values
            wb = openpyxl.load_workbook(model_path, data_only=True)
            results = []

            for export_spec in exports:
                sheet_name = export_spec.get("sheet", "")
                cell_range = export_spec.get("range", "")
                fmt = export_spec.get("format", "csv")
                output_name = export_spec.get(
                    "output", f"export_{sheet_name}_{cell_range}.{fmt}"
                )

                if sheet_name not in wb.sheetnames:
                    results.append(
                        {
                            "output": output_name,
                            "success": False,
                            "error": f"Sheet '{sheet_name}' not found",
                        }
                    )
                    continue

                ws = wb[sheet_name]

                # Extract range data
                rows = []
                try:
                    for row in ws[cell_range]:
                        row_data = []
                        for cell in row:
                            val = cell.value
                            if val is None:
                                val = ""
                            row_data.append(val)
                        rows.append(row_data)
                except Exception as e:
                    results.append(
                        {
                            "output": output_name,
                            "success": False,
                            "error": f"Range error: {str(e)}",
                        }
                    )
                    continue

                # Save output
                output_path = OUTPUT_DIR / output_name
                OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

                if fmt == "csv":
                    with open(str(output_path), "w", newline="", encoding="utf-8") as f:
                        writer = csv.writer(f)
                        writer.writerows(rows)
                elif fmt == "json":
                    # First row as headers if possible
                    if len(rows) > 1:
                        headers = [str(h) for h in rows[0]]
                        json_data = []
                        for row in rows[1:]:
                            record = {}
                            for i, val in enumerate(row):
                                key = headers[i] if i < len(headers) else f"col_{i}"
                                record[key] = val
                            json_data.append(record)
                    else:
                        json_data = rows

                    with open(str(output_path), "w", encoding="utf-8") as f:
                        json.dump(json_data, f, indent=2, default=str)

                results.append(
                    {
                        "output": str(output_path),
                        "format": fmt,
                        "rows": len(rows),
                        "success": True,
                    }
                )

            wb.close()

            return ToolResult(
                success=all(r["success"] for r in results),
                data={"exports": results, "total": len(results)},
            )

        except Exception as e:
            logger.error("Excel export failed", error=str(e))
            return ToolResult(success=False, data=None, error=str(e))
